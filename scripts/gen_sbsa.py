#!/usr/bin/env python3
# Generates supabase/load_sbsa.sql — Standard Bank SD (BDSP) programme update onto the EXISTING sponsor.
# 4 new beneficiaries loaded fresh (CTE + gen_random_uuid); 2 existing (Hearts & Flowers, Leras Events)
# updated in place + missing Annexure-B line added + monthly updates; 2 open escalations. No hand-typed row UUIDs.
import openpyxl, datetime
from openpyxl.utils import column_index_from_string as cidx

SRC="/root/.claude/uploads/8b106a24-0970-5686-987d-486b630f31a1/3f23e435-SBSA_SD_Programme_BDSP_Tracker_final.xlsx"
OUT="/root/uca/supabase/load_sbsa.sql"

SPONSOR="4e401d68-a1e5-4b74-9a67-1ce9511569e4"   # Standard Bank (existing, standalone)
OWNER  ="8957f6bb-9b23-4412-9c14-d25d3e2088de"   # Boitumelo Matobela (manco) — PM/owner (matches existing rows)
RAISER ="69effdfb-6c59-4138-94a4-ea037030a1ec"   # Rinaldo Josie (manco) — escalation raiser
HF_ID  ="ffef76a5-090f-4dc5-9f25-b0ebb788d619"   # existing Hearts & Flowers
LERAS_ID="e2ecdd95-f779-4a45-a0e5-b1ee93a18f7e"  # existing Leras Events
DEF="2026-07-31"

wb=openpyxl.load_workbook(SRC, data_only=True)
def gv(r,col):
    v=r[cidx(col)-1]; return None if v is None else v
def s(v):
    return None if v is None else str(v).strip()
def q(v):
    if v is None: return "null"
    return "'"+str(v).replace("'","''")+"'"
def qd(v):
    if v is None or v=="": return "null"
    if isinstance(v,(datetime.datetime,datetime.date)): return "'"+v.strftime("%Y-%m-%d")+"'"
    t=str(v).strip()
    if t.startswith("00:00") or not t: return "null"
    return "'"+t.replace("'","''")+"'"
def num(v):
    if v is None or v=="": return "null"
    try: return str(round(float(v),2))
    except: return "null"
def money(v):
    try:
        f=float(v)
        return f"R{f:,.0f}" if abs(f-round(f))<0.005 else f"R{f:,.2f}"
    except: return "R0"
def clean_reason(t):
    if t is None: return None
    t=t.strip().lstrip('.').strip()
    # tidy a couple of raw tracker typos in user-facing RAG reasons
    t=t.replace("Engagement status none enaged for july","Engagement recorded as non-engaged for July")
    return t or None

# ---- cleaned per-beneficiary overrides ----
META={
 "SB-001":dict(name="VB Shopfitters", ind="Shopfitting & Retail Interiors", contact="Suveer Bridglall (Manager)", email="ops@vbshopfitters.com", phone="078 351 4734", rag="red", created="2026-01-16", last=None),
 "SB-002":dict(name="Mandla Lighting", ind="Electrical & Lighting", contact="Jeremia (Jerry) Mandla Mpofu", email="jerry@mandlalighting.co.za", phone="011 873 3982", rag="red", created=DEF, last=None),
 "SB-004":dict(name="Sintra Creative", ind="Strategic Communications & Content", contact="Sunilkumar Dulputram Gopal", email="sunil@23x.co.za", phone="072 901 3421", rag="green", created="2026-03-27", last="2026-07-29"),
 "SB-006":dict(name="S&K Panel Beaters", ind="Panel Beating & Auto Repair", contact="King Mfingwana", email="King@skpanelbeaters.co.za", phone="066 189 6398", rag="red", created=DEF, last=None),
}
STATUS={"On Hold":"on_hold","In Progress":"in_progress","Delivered":"completed","Not Started":"not_started","Ongoing":"in_progress"}

# ---- read beneficiary master (rag reason=AY, outstanding=Z, budget=AG) ----
BEN={}
for r in wb["Beneficiary Master"].iter_rows(min_row=5,max_row=10,values_only=True):
    sid=s(gv(r,'A'))
    if not sid: continue
    BEN[sid]=dict(blockers=s(gv(r,'AY')), outdocs=s(gv(r,'Z')), agval=gv(r,'AG'), stage=s(gv(r,'AW')), notes=s(gv(r,'BB')))

# ---- intervention lines grouped by beneficiary id ----
IV={}
for r in wb["Intervention Lines"].iter_rows(min_row=4,max_row=23,values_only=True):
    sid=s(gv(r,'C')); area=s(gv(r,'D'))
    if not sid or not area or not sid.startswith("SB"): continue
    IV.setdefault(sid,[]).append(dict(area=area, acts=s(gv(r,'E')), fee=gv(r,'F'), st=s(gv(r,'G')),
                                      done=gv(r,'H'), note=s(gv(r,'K'))))

# ---- status log grouped by beneficiary id (chronological) ----
STAT={}
for r in wb["Status Log"].iter_rows(min_row=4,max_row=12,values_only=True):
    sid=s(gv(r,'D'))
    if not sid: continue
    STAT.setdefault(sid,[]).append(dict(date=gv(r,'B'), note=s(gv(r,'H')), na=s(gv(r,'I')), owner=s(gv(r,'J'))))

# ---- escalation & risk log ----
ESC={}
for r in wb["Escalation & Risk Log"].iter_rows(min_row=4,max_row=7,values_only=True):
    sid=s(gv(r,'B'))
    if not sid: continue
    ESC[sid]=dict(typ=s(gv(r,'D')), cat=s(gv(r,'E')), desc=s(gv(r,'F')), rating=s(gv(r,'G')),
                  mit=s(gv(r,'H')), date=gv(r,'I'))

def iv_motivation(l):
    m=f"Category: Supplier Development (Annexure B). Fee ex-VAT {money(l['fee'])}."
    if l['acts']: m+=" "+l['acts']
    if l['note']: m+=" "+l['note']
    return m

L=[]
L.append("-- load_sbsa.sql  ·  Standard Bank SD (BDSP) programme — UPDATE onto existing sponsor 'Standard Bank ' ("+SPONSOR+").")
L.append("-- 4 new beneficiaries (VB Shopfitters, Mandla Lighting, Sintra Creative, S&K Panel Beaters) loaded fresh.")
L.append("-- 2 existing (Hearts & Flowers, Leras Events) reconciled in place: reopened to active, RAG/notes refreshed,")
L.append("-- missing Annexure-B line added, monthly Status-Log updates appended. 2 open escalations raised with SBSA ESD.")
L.append("-- Owner/PM: Boitumelo Matobela; escalations raised by Rinaldo Josie. CTE + gen_random_uuid (no literal row UUIDs).")
L.append("begin;")
L.append("")

def emit_new(sid):
    m=META[sid]; b=BEN[sid]; lines=IV.get(sid,[]); st=STAT.get(sid,[])
    rag=m["rag"]; reason = clean_reason(b["blockers"]) if rag!="green" else None
    L.append(f"-- {sid}  {m['name']}  [{b['stage']} -> implementation, {rag}, {len(lines)} lines]")
    L.append("with b as (")
    L.append("  insert into beneficiaries (name,sponsor_id,industry,contact_person,contact_email,contact_phone,directors,"
             "stage,lifecycle,project_manager_id,budget,outstanding_items,rag_override,rag_override_reason,cycle,last_engagement_at,created_at) values (")
    L.append(f"    {q(m['name'])},'{SPONSOR}',{q(m['ind'])},{q(m['contact'])},{q(m['email'])},{q(m['phone'])},'[]'::jsonb,"
             f"'implementation','active','{OWNER}',{num(b['agval'])},{q(b['outdocs'])},'{rag}'::rag,{q(reason)},1,{qd(m['last'])},{qd(m['created'])}) returning id)")
    # interventions — first line captured for weekly-update attach
    for idx,l in enumerate(lines):
        cst=STATUS.get(l['st'],'not_started')
        comp = qd(l['done']) if cst=='completed' else 'null'
        hold = q(l['note'] or b['blockers']) if cst=='on_hold' else 'null'
        alias=f"iv{idx}"
        ret=" returning id" if idx==0 else ""
        L.append(f", {alias} as (")
        L.append("  insert into interventions (beneficiary_id,kind,custom_name,custom_kind,custom_budget,custom_motivation,consultant_id,"
                 "status,completed_at,hold_reason,discovery_status,discovery_at,acknowledged,acknowledged_at,assigned_at,cycle)")
        L.append(f"  select id,'custom',{q(l['area'])},'other',{num(l['fee'])},{q(iv_motivation(l))},'{OWNER}',"
                 f"'{cst}',{comp},{hold},'na',{qd(m['created'])},true,now(),{qd(m['created'])},1 from b{ret})")
    # weekly updates from status log -> attach to first intervention (iv0)
    esc=ESC.get(sid)
    if st:
        # build a UNION insert of all status rows onto iv0
        L.append(", wu as (")
        L.append("  insert into weekly_updates (intervention_id,author_id,completed_work,next_action,blocker_owner,created_at)")
        sels=[]
        for u in st:
            sels.append(f"  select id,'{OWNER}'::uuid,{q(u['note'])},{q(u['na'])},{q(u['owner'])},{qd(u['date'])}::timestamptz from iv0")
        L.append("\n  union all\n".join(sels)+ (" returning 1)" if esc else " returning 1)"))
    if esc:
        eraise = qd(esc['date']) if esc['date'] else qd(DEF)
        reason_txt=f"{esc['typ']}: {esc['cat']}"
        ctx=(esc['desc'] or "")
        if esc['mit']: ctx+=f" Mitigation: {esc['mit']}"
        if esc['rating']: ctx=f"[{esc['rating']} risk] "+ctx
        L.append(", e as (")
        L.append("  insert into escalations (intervention_id,beneficiary_id,reason,context,status,current_owner_id,current_owner_role,"
                 "consultant_id,manco_id,sponsor_id,participants,raised_by,raised_at,last_action_at)")
        L.append(f"  select null,b.id,{q(reason_txt)},{q(ctx)},'with_sponsor',null,'external','{OWNER}','{RAISER}',null,"
                 f"array['{OWNER}','{RAISER}']::uuid[],'{RAISER}',{eraise},{eraise} from b returning id)")
        L.append("insert into escalation_events (escalation_id,at,user_id,kind,to_status,text)")
        L.append(f"  select id,{eraise},'{RAISER}','raised','with_sponsor',{q(reason_txt+': '+ctx)} from e;")
    else:
        # terminate the CTE with a trivial top-level statement referencing wu/iv0
        if st:
            L.append("select 1;")
        else:
            L.append("select 1;")
    L.append("")

for sid in ["SB-001","SB-002","SB-004","SB-006"]:
    emit_new(sid)

# ---- reconcile existing two ----
def reconcile(sid, bid, missing_area, missing_fee, missing_status, missing_acts):
    b=BEN[sid]; m_rag = {"SB-003":"amber","SB-005":"red"}[sid]
    reason=clean_reason(b['blockers'])
    L.append(f"-- {sid}  {'Hearts & Flowers' if sid=='SB-003' else 'Leras Events'}  [reconcile in place: reopen + add line + updates]")
    L.append("update beneficiaries set lifecycle='active', stage='implementation', "
             f"rag_override='{m_rag}'::rag, rag_override_reason={q(reason)}, outstanding_items={q(b['outdocs'])} "
             f"where id='{bid}';")
    # add the missing Annexure-B line (idempotent guard) + attach status-log updates to it
    st=STAT.get(sid,[])
    mot=f"Category: Supplier Development (Annexure B). Fee ex-VAT {money(missing_fee)}. {missing_acts}"
    cst=STATUS.get(missing_status,'in_progress')
    L.append("with iv as (")
    L.append("  insert into interventions (beneficiary_id,kind,custom_name,custom_kind,custom_budget,custom_motivation,consultant_id,"
             "status,discovery_status,discovery_at,acknowledged,acknowledged_at,assigned_at,rag_override,rag_override_reason,cycle)")
    L.append(f"  select '{bid}','custom',{q(missing_area)},'other',{num(missing_fee)},{q(mot)},'{OWNER}',"
             f"'{cst}','na','{DEF}',true,now(),'{DEF}','{m_rag}'::rag,{q(reason)},1")
    L.append(f"  where not exists (select 1 from interventions where beneficiary_id='{bid}' and custom_name={q(missing_area)}) returning id)")
    if st:
        sels=[]
        for u in st:
            sels.append(f"  select id,'{OWNER}'::uuid,{q(u['note'])},{q(u['na'])},{q(u['owner'])},{qd(u['date'])}::timestamptz from iv")
        L.append("insert into weekly_updates (intervention_id,author_id,completed_work,next_action,blocker_owner,created_at)")
        L.append("\n  union all\n".join(sels)+";")
    else:
        L.append("select 1;")
    L.append("")

reconcile("SB-003", HF_ID, "Strategy and Brand Audit", 65384, "In Progress",
          "Strategy and brand audit forming part of the revised marketing-led scope agreed with the beneficiary in December 2025. COMBINED R65,384 value covers this line and the delivered Business Profile; the split is not documented. No signed Annexure B variation exists (flag F28).")
reconcile("SB-005", LERAS_ID, "Monitoring, Evaluation and Reporting (Ember360)", 0, "Ongoing",
          "Continuous Ember360 monitoring line — stated as 'Included' in Annexure B at no separate fee. Runs for the life of the engagement; never marked Delivered.")

L.append("select (select count(*) from beneficiaries where sponsor_id='"+SPONSOR+"') bens,"
         "(select count(*) from interventions i join beneficiaries b on b.id=i.beneficiary_id where b.sponsor_id='"+SPONSOR+"') ivs,"
         "(select count(*) from escalations e join beneficiaries b on b.id=e.beneficiary_id where b.sponsor_id='"+SPONSOR+"') escs,"
         "(select count(*) from weekly_updates wu join interventions i on i.id=wu.intervention_id join beneficiaries b on b.id=i.beneficiary_id where b.sponsor_id='"+SPONSOR+"') wus;")
L.append("commit;")
L.append("")
L.append("-- ROLLBACK NOTE: this load both INSERTS (4 new benes) and UPDATES 2 existing benes in place.")
L.append("-- To reverse the new beneficiaries only (VB Shopfitters, Mandla Lighting, Sintra Creative, S&K Panel Beaters):")
L.append("--   delete from escalation_events where escalation_id in (select e.id from escalations e join beneficiaries b on b.id=e.beneficiary_id where b.sponsor_id='"+SPONSOR+"' and b.name in ('VB Shopfitters','Mandla Lighting','Sintra Creative','S&K Panel Beaters'));")
L.append("--   delete from escalations where beneficiary_id in (select id from beneficiaries where sponsor_id='"+SPONSOR+"' and name in ('VB Shopfitters','Mandla Lighting','Sintra Creative','S&K Panel Beaters'));")
L.append("--   delete from weekly_updates where intervention_id in (select i.id from interventions i join beneficiaries b on b.id=i.beneficiary_id where b.sponsor_id='"+SPONSOR+"' and b.name in ('VB Shopfitters','Mandla Lighting','Sintra Creative','S&K Panel Beaters'));")
L.append("--   delete from interventions where beneficiary_id in (select id from beneficiaries where sponsor_id='"+SPONSOR+"' and name in ('VB Shopfitters','Mandla Lighting','Sintra Creative','S&K Panel Beaters'));")
L.append("--   delete from beneficiaries where sponsor_id='"+SPONSOR+"' and name in ('VB Shopfitters','Mandla Lighting','Sintra Creative','S&K Panel Beaters');")
L.append("-- The 2 reconciled benes' added lines: delete from interventions where beneficiary_id in ('"+HF_ID+"','"+LERAS_ID+"') and custom_name in ('Strategy and Brand Audit','Monitoring, Evaluation and Reporting (Ember360)');")

open(OUT,"w").write("\n".join(L))
print("WROTE",OUT,"lines:",len(L))
print("new benes:4  reconciled:2  escalations:",len(ESC),"  status rows:",sum(len(v) for v in STAT.values()))
