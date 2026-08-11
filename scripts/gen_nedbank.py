#!/usr/bin/env python3
# Generates supabase/load_nedbank.sql for the Nedbank SIU SMME Investment Readiness Programme.
# CTE-per-SMME with gen_random_uuid() (DB-generated ids) — no hand-typed row UUIDs. 15 beneficiaries.
import openpyxl, datetime
from openpyxl.utils import column_index_from_string as cidx

SRC="/root/.claude/uploads/8b106a24-0970-5686-987d-486b630f31a1/3efff30a-Nedbank_SMME_Programme_Tracker_3.xlsx"
OUT="/root/uca/supabase/load_nedbank.sql"

SPONSOR="28d8ffac-d2fb-4ddb-acd7-86c580c23091"   # Nedbank SIU
OWNER  ="8957f6bb-9b23-4412-9c14-d25d3e2088de"   # Boitumelo Matobela (manco) - record owner + update author
RAISER ="69effdfb-6c59-4138-94a4-ea037030a1ec"   # Rinaldo Josie (manco) - escalation raiser
DEFDATE="2026-08-05"

wb=openpyxl.load_workbook(SRC, data_only=True)
def gv(r,col):
    v=r[cidx(col)-1]; return None if v is None else v
def s(v):
    return None if v is None else str(v).strip()

dd={}
for r in wb["DD Baseline"].iter_rows(min_row=4, max_row=18, values_only=True):
    sid=s(r[0])
    if not sid or not sid.startswith("NED"): continue
    dd[sid]=(r[cidx('L')-1], s(r[cidx('N')-1]), s(r[cidx('O')-1]), r[cidx('C')-1])
stat={}
for r in wb["Status Log"].iter_rows(min_row=4, max_row=17, values_only=True):
    sid=s(r[cidx('C')-1])
    if not sid: continue
    stat[sid]=(r[cidx('B')-1], s(r[cidx('G')-1]), s(r[cidx('H')-1]), s(r[cidx('I')-1]))
esc={}
for r in wb["Escalation Log"].iter_rows(min_row=4, max_row=12, values_only=True):
    sid=s(r[cidx('B')-1])
    if not sid: continue
    esc[sid]=(s(r[cidx('F')-1]), s(r[cidx('G')-1]))
esc["NED-012"]=("Non-Delivery",
    "Assessed ELIGIBLE at DD (83%, 15 Apr 2026), one of the two highest scores in the cohort, but no coaching delivery on record. "
    "DD filed under the entity name Rokiwaste (Pty) Ltd; Wasteq is the cohort/trading name. In-person meeting being scheduled.")

def q(v):
    if v is None: return "null"
    return "'"+str(v).replace("'","''")+"'"
def qd(v):
    if v is None or v=="": return "null"
    if isinstance(v,(datetime.datetime,datetime.date)): return "'"+v.strftime("%Y-%m-%d")+"'"
    txt=str(v).strip()
    if txt.startswith("00:00") or not txt: return "null"
    if "/" in txt:
        d,m,y=txt.split("/"); return f"'{int(y):04d}-{int(m):02d}-{int(d):02d}'"
    return "'"+txt.replace("'","''")+"'"

L=[]
L.append("-- load_nedbank.sql  ·  Nedbank SIU SMME Investment Readiness Programme (sponsor 'Nedbank SIU').")
L.append("-- CTE-per-SMME, DB-generated UUIDs. 15 beneficiaries (Ember360+DD complete); each: DD intervention (completed)")
L.append("-- + 6-session coaching intervention. Owner: Boitumelo Matobela (manco); coaches preserved in history.")
L.append("-- 10 open escalations raised with Nedbank SIU (9 Escalation Log + NED-012 master flag). Stage: Tranche 1 Complete")
L.append("-- -> monitoring, else implementation. RAG pinned to the tracker via overrides.")
L.append("begin;")
L.append("")

for r in wb["SMME Master"].iter_rows(min_row=5, max_row=19, values_only=True):
    sid=s(gv(r,'A'))
    if not sid or not sid.startswith("NED"): continue
    name=s(gv(r,'B')); sector=s(gv(r,'C')); prov=s(gv(r,'D')); coach=s(gv(r,'E'))
    docstat=s(gv(r,'K')); outdocs=s(gv(r,'L')); bankN=gv(r,'N')
    U=int(gv(r,'U') or 0); W=s(gv(r,'W')); X=gv(r,'X')
    tstage=s(gv(r,'AG')); rag=(s(gv(r,'AH')) or "amber").lower()
    blockers=s(gv(r,'AI')); notes=s(gv(r,'AL'))
    ddrp,ddelig,ddweak,ddd = dd.get(sid,(None,None,None,None))
    stg = "monitoring" if tstage=="Tranche 1 Complete" else "implementation"
    cst = "completed" if U>=6 else ("in_progress" if U>=1 else ("awaiting_beneficiary" if tstage=="Escalated" else "not_started"))
    created = qd(ddd) if qd(ddd)!="null" else qd(DEFDATE)
    ddlit   = created
    last_eng= qd(X) if X else (qd(stat.get(sid,(None,))[0]) if stat.get(sid) else "null")
    if last_eng=="null": last_eng=created

    iva_mot=f"Category: Diagnostics & Due Diligence. Baseline DD readiness {ddrp if ddrp is not None else 'n/a'} ({ddelig or 'n/a'})."
    if ddweak: iva_mot+=f" Weakest at baseline: {ddweak}."
    if prov: iva_mot+=f" Province: {prov}."
    if docstat: iva_mot+=f" Compliance: {docstat}."
    ivb_mot=f"Category: Coaching & Mentorship. Coach: {coach or 'UCH'}. {U} of 6 sessions delivered (F1-F4, BS1, BS2); booklets {W or 'n/a'}."
    if bankN is not None: ivb_mot+=f" Bankability {bankN}/10."
    if notes: ivb_mot+=f" {notes}"
    completed_at = qd(X) if cst=="completed" else "null"
    ivb_over = f"'{rag}'::rag" if cst!="completed" else "null"
    ivb_over_reason = q(blockers) if cst!="completed" else "null"
    aw = qd(DEFDATE) if cst=="awaiting_beneficiary" else "null"

    st=stat.get(sid)
    if st:
        wd,wnote,wna,wown=st
        wu_cw=q(wnote); wu_ip="null"; wu_na=q(wna); wu_bo=q(wown); wu_at=qd(wd) if wd else qd(DEFDATE)
    else:
        wu_cw="null"; wu_ip=q(notes or blockers); wu_na=q("Schedule first coaching session and confirm coach capacity"); wu_bo="null"; wu_at=qd(DEFDATE)

    has_esc = sid in esc
    eraise = qd(X) if X else qd(DEFDATE)

    L.append(f"-- {sid}  {name}  [{tstage} -> {stg}, {rag}, {U}/6]")
    L.append("with b as (")
    L.append("  insert into beneficiaries (name,sponsor_id,industry,directors,stage,lifecycle,project_manager_id,"
             "outstanding_items,rag_override,rag_override_reason,cycle,last_engagement_at,created_at) values (")
    L.append(f"    {q(name)},'{SPONSOR}',{q(sector)},'[]'::jsonb,'{stg}','active','{OWNER}',"
             f"{q(outdocs)},'{rag}'::rag,{q(blockers)},1,{last_eng},{created}) returning id),")
    L.append("iva as (")
    L.append("  insert into interventions (beneficiary_id,kind,custom_name,custom_kind,custom_motivation,consultant_id,"
             "status,completed_at,discovery_status,discovery_at,acknowledged,acknowledged_at,assigned_at,cycle)")
    L.append(f"  select b.id,'custom','Ember360 diagnostic & Due Diligence assessment','other',{q(iva_mot)},'{OWNER}',"
             f"'completed',{ddlit},'na',{ddlit},true,now(),{ddlit},1 from b),")
    L.append("ivb as (")
    L.append("  insert into interventions (beneficiary_id,kind,custom_name,custom_kind,custom_motivation,consultant_id,"
             "status,completed_at,awaiting_response_since,discovery_status,discovery_at,acknowledged,acknowledged_at,"
             "assigned_at,rag_override,rag_override_reason,cycle)")
    L.append(f"  select b.id,'custom','Investment-readiness coaching — 6 sessions (F1-F4, BS1, BS2)','other',{q(ivb_mot)},'{OWNER}',"
             f"'{cst}',{completed_at},{aw},'na',{ddlit},true,now(),{ddlit},{ivb_over},{ivb_over_reason},1 from b returning id)")
    if not has_esc:
        L.append("insert into weekly_updates (intervention_id,author_id,completed_work,in_progress,next_action,blocker_owner,created_at)")
        L.append(f"  select id,'{OWNER}',{wu_cw},{wu_ip},{wu_na},{wu_bo},{wu_at} from ivb;")
    else:
        cat,detail=esc[sid]
        L.append(", wu as (")
        L.append("  insert into weekly_updates (intervention_id,author_id,completed_work,in_progress,next_action,blocker_owner,created_at)")
        L.append(f"  select id,'{OWNER}',{wu_cw},{wu_ip},{wu_na},{wu_bo},{wu_at} from ivb returning 1),")
        L.append("e as (")
        L.append("  insert into escalations (intervention_id,beneficiary_id,reason,context,status,current_owner_id,current_owner_role,"
                 "consultant_id,manco_id,sponsor_id,participants,raised_by,raised_at,last_action_at)")
        L.append(f"  select ivb.id,b.id,{q(cat)},{q(detail)},'with_sponsor',null,'external','{OWNER}','{RAISER}',null,"
                 f"array['{OWNER}','{RAISER}']::uuid[],'{RAISER}',{eraise},{eraise} from ivb,b returning id)")
        L.append("insert into escalation_events (escalation_id,at,user_id,kind,to_status,text)")
        L.append(f"  select id,{eraise},'{RAISER}','raised','with_sponsor',{q((cat or '')+': '+(detail or ''))} from e;")
    L.append("")

L.append("select (select count(*) from beneficiaries where sponsor_id='"+SPONSOR+"') bens,"
         "(select count(*) from interventions i join beneficiaries b on b.id=i.beneficiary_id where b.sponsor_id='"+SPONSOR+"') ivs,"
         "(select count(*) from escalations e join beneficiaries b on b.id=e.beneficiary_id where b.sponsor_id='"+SPONSOR+"') escs,"
         "(select count(*) from weekly_updates wu join interventions i on i.id=wu.intervention_id join beneficiaries b on b.id=i.beneficiary_id where b.sponsor_id='"+SPONSOR+"') wus;")
L.append("commit;")
L.append("")
L.append("-- ROLLBACK (Nedbank SIU only):")
L.append("-- delete from escalation_events where escalation_id in (select e.id from escalations e join beneficiaries b on b.id=e.beneficiary_id where b.sponsor_id='"+SPONSOR+"');")
L.append("-- delete from escalations where beneficiary_id in (select id from beneficiaries where sponsor_id='"+SPONSOR+"');")
L.append("-- delete from weekly_updates where intervention_id in (select i.id from interventions i join beneficiaries b on b.id=i.beneficiary_id where b.sponsor_id='"+SPONSOR+"');")
L.append("-- delete from interventions where beneficiary_id in (select id from beneficiaries where sponsor_id='"+SPONSOR+"');")
L.append("-- delete from beneficiaries where sponsor_id='"+SPONSOR+"';")

open(OUT,"w").write("\n".join(L))
print("WROTE",OUT,"lines:",len(L),"  escalated SMMEs:",len(esc))
